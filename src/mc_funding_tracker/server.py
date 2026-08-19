"""Flask web dashboard for mc-funding-tracker.

Research (a web-search API call) runs on a background thread tracked via jobs.py
rather than blocking the request — it can take up to ~180s, which was long
enough to hit real client-side timeouts when it ran synchronously. There's no
native GUI here (unlike meetcap's menu bar app), so background threads have no
thread-safety concerns beyond the job tracker's own lock.
"""
from __future__ import annotations

import datetime
import io
import logging
import os
import threading
import webbrowser
from typing import Any, Dict

from flask import Flask, flash, redirect, render_template, request, send_file, url_for
from openpyxl import Workbook
from openpyxl.styles import Font
from werkzeug.serving import make_server

from . import db, jobs
from .research import parse_funding_update, run_research

logger = logging.getLogger(__name__)

PORT = 5430

_server = None
_thread = None


def _format_usd(amount) -> str:
    """Render a dollar amount compactly, e.g. 2_000_000 -> '$2M'."""
    if amount is None:
        return "undisclosed"
    if amount >= 1_000_000:
        return f"${amount / 1_000_000:.1f}M".replace(".0M", "M")
    if amount >= 1_000:
        return f"${amount / 1_000:.0f}K"
    return f"${amount}"


def _format_class_year(year) -> str:
    """Render a class year compactly, e.g. 2018 -> "'18"."""
    if not year:
        return ""
    return f"'{str(int(year))[-2:]}"


def _filter_and_sort_companies():
    """Apply the grad_years/exited filters and sort from query params. Shared by the
    index page and the Excel export so the export actually matches what's on screen."""
    # Default-direction-per-column (name starts ascending, total_funding starts
    # descending — highest funding first is the useful default for a money column)
    # so the first click on a header does something sensible before toggling.
    default_dirs = {"name": "asc", "total_funding": "desc"}
    sort = request.args.get("sort", "name")
    if sort not in default_dirs:
        sort = "name"
    direction = request.args.get("dir", default_dirs[sort])
    if direction not in ("asc", "desc"):
        direction = default_dirs[sort]

    grad_years = request.args.get("grad_years", "")
    cutoff_year = None
    if grad_years.isdigit():
        cutoff_year = datetime.date.today().year - int(grad_years)

    exited_only = request.args.get("exited") == "on"

    funding_status = request.args.get("funding_status", "confirmed")
    if funding_status not in ("confirmed", "all"):
        funding_status = "confirmed"

    companies = db.get_companies(include_unconfirmed=(funding_status == "all"))
    if cutoff_year is not None:
        companies = [
            c for c in companies
            if any(f["class_year"] and f["class_year"] >= cutoff_year for f in c["founders"])
        ]
    if exited_only:
        exited_ids = db.get_exited_company_ids()
        companies = [c for c in companies if c["id"] in exited_ids]

    if sort == "name":
        companies.sort(key=lambda c: c["name"].lower(), reverse=(direction == "desc"))
    else:
        companies.sort(key=lambda c: c["total_funding"], reverse=(direction == "desc"))

    next_dir = {col: ("desc" if (sort == col and direction == "asc") else "asc" if (sort == col and direction == "desc") else default_dirs[col]) for col in default_dirs}

    return companies, sort, direction, next_dir, grad_years, exited_only, funding_status


def _companies_to_xlsx(companies) -> io.BytesIO:
    """Build an .xlsx workbook mirroring the company list table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"

    headers = [
        "Company", "Founder(s)", "Dartmouth IP", "Total Funding",
        "Last Round Type", "Last Round Amount", "Last Round Status",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for c in companies:
        founders = ", ".join(
            f["name"] + (f" '{str(f['class_year'])[-2:]}" if f["class_year"] else "")
            for f in c["founders"]
        )
        latest = c["latest_round"]
        if latest:
            round_type, amount, status = latest["round_type"], latest["amount_usd"], latest["status"]
        elif c["last_researched_at"]:
            round_type, amount, status = "no rounds found", None, ""
        else:
            round_type, amount, status = "no research", None, ""
        ws.append([
            c["name"], founders, "Yes" if c["dartmouth_ip"] else "",
            c["total_funding"], round_type, amount, status,
        ])

    money_format = '"$"#,##0'
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        row[0].number_format = money_format
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        row[0].number_format = money_format

    widths = [28, 32, 12, 14, 24, 16, 14]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def create_app(config: Dict[str, Any]) -> Flask:
    """Create and configure the Flask application."""
    app = Flask(__name__)
    app.secret_key = os.urandom(24)
    app.jinja_env.filters["usd"] = _format_usd
    app.jinja_env.filters["classyear"] = _format_class_year
    db.init_db()

    @app.route("/")
    def index():
        companies, sort, direction, next_dir, grad_years, exited_only, funding_status = _filter_and_sort_companies()

        total_funding = sum(c["total_funding"] for c in companies)
        researching_ids = jobs.running_ids()
        return render_template(
            "index.html",
            companies=companies,
            total_funding=total_funding,
            researching_ids=researching_ids,
            sort=sort,
            direction=direction,
            next_dir=next_dir,
            grad_years=grad_years,
            exited_only=exited_only,
            funding_status=funding_status,
        )

    @app.route("/export.xlsx")
    def export_companies():
        companies, *_rest = _filter_and_sort_companies()
        buf = _companies_to_xlsx(companies)
        filename = f"mc-funding-tracker-{datetime.date.today().isoformat()}.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    @app.route("/company/new")
    def new_company():
        return render_template("add_company.html")

    @app.route("/company", methods=["POST"])
    def create_company():
        name = request.form.get("name", "").strip()
        website = request.form.get("website", "").strip()
        dartmouth_ip = request.form.get("dartmouth_ip") == "yes"

        founders = []
        for fname, fyear in zip(
            request.form.getlist("founder_name"),
            request.form.getlist("founder_class_year"),
        ):
            fname = fname.strip()
            if not fname:
                continue
            class_year = int(fyear) if fyear.strip().isdigit() else None
            founders.append({"name": fname, "class_year": class_year})

        if not name:
            flash("Company name is required.", "error")
            return redirect(url_for("new_company"))

        company_id = db.add_company(name, website, dartmouth_ip, founders)
        flash(f"Added {name}.", "success")
        return redirect(url_for("company_detail", company_id=company_id))

    @app.route("/company/<int:company_id>")
    def company_detail(company_id: int):
        company = db.get_company(company_id)
        if company is None:
            return "Company not found", 404

        result = jobs.pop_result(company_id)
        if result is not None:
            if result["status"] == "done":
                summary = result["summary"]
                msg = (
                    f"Research complete: web search found {summary['web_found']} "
                    f"({summary['web_inserted']} new)."
                )
                flash(msg, "error" if summary["errors"] else "success")
                for err in summary["errors"]:
                    flash(err, "error")
            else:
                flash(f"Research failed: {result['error']}", "error")

        return render_template(
            "company.html", company=company, researching=jobs.is_running(company_id)
        )

    @app.route("/company/<int:company_id>/research", methods=["POST"])
    def research_company(company_id: int):
        company = db.get_company(company_id)
        if company is None:
            return "Company not found", 404
        started = jobs.start(company_id, lambda: run_research(company_id, config))
        if not started:
            flash("Research is already running for this company.", "error")
        return redirect(url_for("company_detail", company_id=company_id))

    @app.route("/company/<int:company_id>/report-update", methods=["POST"])
    def report_funding_update(company_id: int):
        text = request.form.get("body", "").strip()
        if not text:
            return redirect(url_for("company_detail", company_id=company_id))
        try:
            parsed_rounds = parse_funding_update(text, config)
            if not parsed_rounds:
                flash("Could not find any funding round in that update.", "error")
                return redirect(url_for("company_detail", company_id=company_id))
            for parsed in parsed_rounds:
                db.add_funding_round(
                    company_id=company_id,
                    round_type=parsed.get("round_type"),
                    amount_usd=parsed.get("amount_usd"),
                    announced_date=parsed.get("announced_date"),
                    investors=parsed.get("investors"),
                    source="internal",
                    source_url=None,
                )
            round_types = ", ".join(r.get("round_type") or "funding round" for r in parsed_rounds)
            flash(f"Recorded {len(parsed_rounds)} round(s): {round_types}.", "success")
        except Exception as e:
            logger.exception(f"Failed to parse funding update for company_id={company_id}")
            flash(f"Could not record that update: {e}", "error")
        return redirect(url_for("company_detail", company_id=company_id))

    @app.route("/round/<int:round_id>/confirm", methods=["POST"])
    def confirm_round_route(round_id: int):
        db.confirm_round(round_id)
        return redirect(request.referrer or url_for("index"))

    @app.route("/round/<int:round_id>/reject", methods=["POST"])
    def reject_round_route(round_id: int):
        db.reject_round(round_id)
        flash("Marked as not this company — it won't be re-suggested from that filer.", "success")
        return redirect(request.referrer or url_for("index"))

    @app.route("/round/<int:round_id>/unreject", methods=["POST"])
    def unreject_round_route(round_id: int):
        db.unreject_round(round_id)
        return redirect(request.referrer or url_for("index"))

    @app.route("/round/<int:round_id>/unconfirm", methods=["POST"])
    def unconfirm_round_route(round_id: int):
        db.unconfirm_round(round_id)
        flash("Round moved back to unconfirmed.", "success")
        return redirect(request.referrer or url_for("index"))

    @app.route("/company/<int:company_id>/edit", methods=["POST"])
    def edit_company_route(company_id: int):
        name = request.form.get("name", "").strip()
        website = request.form.get("website", "").strip()
        dartmouth_ip = request.form.get("dartmouth_ip") == "yes"
        if not name:
            flash("Company name is required.", "error")
            return redirect(url_for("company_detail", company_id=company_id))
        db.update_company(company_id, name, website, dartmouth_ip)
        flash("Company updated.", "success")
        return redirect(url_for("company_detail", company_id=company_id))

    @app.route("/company/<int:company_id>/founder", methods=["POST"])
    def add_founder_route(company_id: int):
        name = request.form.get("name", "").strip()
        class_year_raw = request.form.get("class_year", "").strip()
        class_year = int(class_year_raw) if class_year_raw.isdigit() else None
        if not name:
            flash("Founder name is required.", "error")
            return redirect(url_for("company_detail", company_id=company_id))
        db.add_founder(company_id, name, class_year)
        flash(f"Added {name} as a founder.", "success")
        return redirect(url_for("company_detail", company_id=company_id))

    @app.route("/founder/<int:founder_id>/edit", methods=["POST"])
    def edit_founder_route(founder_id: int):
        name = request.form.get("name", "").strip()
        class_year_raw = request.form.get("class_year", "").strip()
        class_year = int(class_year_raw) if class_year_raw.isdigit() else None
        if not name:
            flash("Founder name is required.", "error")
            return redirect(request.referrer or url_for("index"))
        db.update_founder(founder_id, name, class_year)
        flash("Founder updated.", "success")
        return redirect(request.referrer or url_for("index"))

    @app.route("/founder/<int:founder_id>/delete", methods=["POST"])
    def delete_founder_route(founder_id: int):
        db.delete_founder(founder_id)
        flash("Founder deleted.", "success")
        return redirect(request.referrer or url_for("index"))

    @app.route("/round/<int:round_id>/edit", methods=["POST"])
    def edit_round_route(round_id: int):
        round_type = request.form.get("round_type", "").strip() or None
        amount_raw = request.form.get("amount_usd", "").strip()
        amount_usd = int(amount_raw) if amount_raw.isdigit() else None
        announced_date = request.form.get("announced_date", "").strip() or None
        investors = request.form.get("investors", "").strip() or None
        db.update_funding_round(round_id, round_type, amount_usd, announced_date, investors)
        flash("Funding round updated.", "success")
        return redirect(request.referrer or url_for("index"))

    return app


def start_server(config: Dict[str, Any]) -> str:
    """Start the Werkzeug server in a daemon thread. Returns the base URL."""
    global _server, _thread
    if _server is not None:
        return f"http://127.0.0.1:{PORT}"

    app = create_app(config)
    _server = make_server("127.0.0.1", PORT, app, threaded=True)
    _thread = threading.Thread(target=_server.serve_forever, daemon=True)
    _thread.start()
    return f"http://127.0.0.1:{PORT}"


def open_dashboard(config: Dict[str, Any]) -> None:
    """Start the server (if needed) and open the dashboard in a browser."""
    url = start_server(config)
    webbrowser.open(url)
